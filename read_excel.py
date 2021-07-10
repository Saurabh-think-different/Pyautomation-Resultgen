from openpyxl import load_workbook, Workbook

def read_excel():
    values = []
    wb = load_workbook(filename = 'regnos.xlsx')
    ws = wb.active
    for row in range(2, 100):
        value = ws['A'+str(row)].value
        if(value != None):
            values.append(value)
    
    return values

def write_excel(data):
    wb = Workbook()
    ws = wb.active
    for record in data:
        record['std_marks'].append(record['regno'])
        ws.append(record['std_marks'])
    wb.save("sample.xlsx")